import os
import pandas as pd
import openpyxl

# Função para ler os arquivos de análise
def read_analysis_files(root_folder, num_folders):
    data = []

    for i in range(1, num_folders + 1):
        folder_name = f"Sudoku_{i}"
        analysis_file_path = os.path.join(root_folder, folder_name, f"analiz_sudoku_{i}.txt")
        if os.path.exists(analysis_file_path):
            with open(analysis_file_path, 'r') as file:
                lines = file.readlines()
                if len(lines) >= 5:
                    execution_time_code = float(lines[0].strip())
                    execution_time_create_sudoku = float(lines[1].strip())
                    execution_time_coloring = float(lines[2].strip())
                    num_recursive_calls = int(lines[3].strip())
                    num_colors_tested = int(lines[4].strip())

                    data.append({
                        'Pasta': folder_name,
                        'Tempo de execucao do codigo': execution_time_code,
                        'Tempo de execucao gerar Sudoku': execution_time_create_sudoku,
                        'Tempo de execucao do algoritmo de coloracao': execution_time_coloring,
                        'Numero de chamadas recursivas': num_recursive_calls,
                        'Numero de cores testadas': num_colors_tested
                    })
                else:
                    print(f"File {analysis_file_path} does not have enough data.")
        else:
            print(f"File {analysis_file_path} not found.")

    return data

# Função para calcular as medianas
def calculate_median_metrics(data):
    df = pd.DataFrame(data)
    df_numeric = df.drop(columns=['Pasta'])  # Remove a coluna 'Pasta'
    sudoku_menor_tempo = df.loc[df['Tempo de execucao do algoritmo de coloracao'].idxmin()]['Pasta']
    sudoku_maior_tempo = df.loc[df['Tempo de execucao do algoritmo de coloracao'].idxmax()]['Pasta']
    # Calcula as medianas apenas para as colunas numéricas
    median_values = df_numeric.median()
    return median_values

def style_table(worksheet, start_col, end_col, start_row, end_row):
    destaque = openpyxl.styles.PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    preenchimento = openpyxl.styles.PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    borda = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin', color='000000'),
                                   right=openpyxl.styles.Side(style='thin', color='000000'),
                                   top=openpyxl.styles.Side(style='thin', color='000000'),
                                   bottom=openpyxl.styles.Side(style='thin', color='000000'))

    for row in range(start_row, end_row + 1):
        worksheet.cell(row=row, column=start_col).fill = destaque
        worksheet.cell(row=row, column=start_col).border = borda

    for col in range(start_col, end_col + 1):
        worksheet.cell(row=start_row, column=col).fill = destaque
        worksheet.cell(row=start_row, column=col).border = borda

    for row in range(start_row, end_row + 1):
        worksheet.cell(row=row, column=start_col).fill = destaque
        worksheet.cell(row=row, column=start_col).border = borda

    for row in range(start_row + 1, end_row + 1):
        for col in range(start_col + 1, end_col + 1):
            worksheet.cell(row=row, column=col).fill = preenchimento
            worksheet.cell(row=row, column=col).border = borda

    for col in range(start_col, end_col + 1):
        max_length = 50
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max_length * 1.2
        
# Carregar dados e calcular as medianas
root_folder = "Output"
num_folders = 100

data = read_analysis_files(root_folder, num_folders)
median_metrics = calculate_median_metrics(data)

# Salvar dados brutos na planilha "Resultados"
excel_file = 'results.xlsx'

book = openpyxl.load_workbook(excel_file)
if 'Resultados' in book.sheetnames:
    sheet = book['Resultados']
    book.remove(sheet)

with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df = pd.DataFrame(data)
    df.to_excel(writer, sheet_name='Resultados', startrow=0, startcol=0, header=True, index=False)
    workbook = writer.book
    worksheet = writer.sheets['Resultados']
    style_table(worksheet,start_col=1, end_col=6, start_row=1, end_row=num_folders+1)  # Estilizar tabela 2
    table = worksheet.tables.values()
    if table:
        table = list(table)[0]
        table.theme = "TableStyleMedium9"  # Definindo um estilo de tabela
    workbook.save(excel_file)

# Adicionar medianas ao arquivo Excel na planilha "Medianas"
# Adicionar medianas ao arquivo Excel na planilha "Medianas"
with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    median_metrics.to_excel(writer, sheet_name='Medianas', startrow=1, startcol=1, header=True, index=True)
    workbook = writer.book
    worksheet = writer.sheets['Medianas']
    table = worksheet.tables.values()
    if table:
        table = list(table)[0]
        table.theme = "TableStyleMedium9"  # Definindo um estilo de tabela

    # Aplicar estilos à tabela de Medianas chamando a função style_excel_table

    # Encontrar o Sudoku com o menor e o maior tempo de execução do algoritmo de coloração
    sudoku_menor_tempo = df.loc[df['Tempo de execucao do algoritmo de coloracao'].idxmin()]['Pasta']
    sudoku_maior_tempo = df.loc[df['Tempo de execucao do algoritmo de coloracao'].idxmax()]['Pasta']

    # Adicionar colunas de melhor e pior caso na planilha "Medianas"
    worksheet.cell(row=2, column=3).value = 'Mediana (' + str(num_folders) + ' sudoku)'  # Atualiza o cabeçalho da coluna B
    worksheet.cell(row=9, column=3).value = 'Tempo de execucao do algoritmo de coloracao'
    worksheet.cell(row=10, column=2).value = 'Melhor Caso'
    worksheet.cell(row=11, column=2).value = 'Pior Caso'
    worksheet.cell(row=10, column=3).value = f'{sudoku_menor_tempo} - {df["Tempo de execucao do algoritmo de coloracao"].min()}'
    worksheet.cell(row=11, column=3).value = f'{sudoku_maior_tempo} - {df["Tempo de execucao do algoritmo de coloracao"].max()}'

    # Alpicar estilizacao tabel 2
    negrito = openpyxl.styles.Font(bold=True)
    centralizado = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    worksheet.cell(row=9, column=3).font = negrito
    worksheet.cell(row=9, column=3).alignment = centralizado
    worksheet.cell(row=10, column=2).font = negrito
    worksheet.cell(row=10, column=2).alignment = centralizado
    worksheet.cell(row=11, column=2).font = negrito
    worksheet.cell(row=11, column=2).alignment = centralizado
    
    style_table(worksheet,start_col=2, end_col=3, start_row=2, end_row=7)  # Estilizar tabela 1
    style_table(worksheet,start_col=2, end_col=3, start_row=9, end_row=11)  # Estilizar tabela 2
    
    workbook.save(excel_file)

print("Dados brutos e medianas das métricas adicionados ao arquivo Excel.")